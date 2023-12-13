import datetime
import os

from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, HttpResponse, Http404
from django.shortcuts import render
from django.template import Library
from django.urls import reverse_lazy
from django.views import generic
from django.views.generic.edit import FormMixin
from openpyxl.workbook import Workbook

from apps.common import importFRC
from apps.entry.forms import MatchScoutForm, ImportForm
from apps.entry.imports import import_first, get_match_data_event
from apps.entry.models import Images, Team, Match, Pits
from apps.organization.models import OrgSettings, Event
from apps.entry.views.helpers import get_present_teams, make_int
from swiss import settings

register = Library

# @user_passes_test(scout_lead_check, login_url='organization:login')
@login_required(login_url='organization:login')
def download(request):
    # TODO find a way to prevent spamming this.
    path = 'match_history.xlsx'
    path = os.path.join(settings.BASE_DIR, path)
    update_csv(request.user.orgmember.organization)

    if request.method == "GET" and os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="text/xlsx")
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(path)
            return response

    return Http404


def update_csv(organization):
    print("Updating XLSX File")

    workbook = Workbook()
    writable_models = [Match, Pits]

    for model in writable_models:
        workbook.create_sheet(title=model._meta.model_name)
        sheet = workbook.get_sheet_by_name(model._meta.model_name)
        headers = model._meta.get_fields()
        x = 1
        for header in headers:
            sheet.cell(column=x, row=1, value=str(header))
            x += 1
        data = model.objects.all().values_list().filter(ownership=organization)
        x = 1
        y = 2
        for entry in data:
            for field in entry:
                sheet.cell(column=x, row=y, value=str(field))
                x += 1
            y += 1

    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    workbook.save("match_history.xlsx")


@login_required(login_url='organization:login')
def write_image_upload(request):
    if request.method == 'POST':
        team_number = make_int(request.POST.get('teamNumber'))
        team = Team.objects.get(number=team_number)

        request.session.set_test_cookie()

        files = request.FILES
        files = files.popitem()[1]

        for file in files:
            file.name = str(team_number) + "-----" + str(datetime.datetime.now()).replace('.', '') + '.jpg'
            image = Images(name=team.name, image=file)
            image.save()
            team.images.add(image)
            team.save()
        return HttpResponseRedirect(reverse_lazy('entry:index'))
    else:
        return HttpResponseRedirect(reverse_lazy('entry:index'))


@login_required(login_url='organization:login')
def import_from_first(request):
    if request.method == 'GET':
        return HttpResponseRedirect(reverse_lazy('entry:import'))

    elif request.method == 'POST':
        import_type = make_int(request.POST.get('importType'))
        key = request.POST.get('key')
        year = request.POST.get('year')

        if year == 0:
            year = None

        if import_type == 0:
            importFRC.import_district(key, year)
        elif import_type == 1:
            importFRC.import_event(key, year)
        elif import_type == 3:
            importFRC.import_team(key, year)

    return HttpResponseRedirect(reverse_lazy('entry:import'))


@login_required(login_url='organization:login')
def import_schedule_from_first(request):
    teamsettings = OrgSettings.objects.all().filter(org_id=request.user.orgmember.organization_id)[0]
    first_key = Event.objects.all().filter(id=make_int(teamsettings.current_event.id))[0].FIRST_key
    importFRC.import_schedule(first_key, playoffs=False)
    importFRC.import_schedule(first_key, playoffs=True)

    return HttpResponseRedirect(reverse_lazy('entry:schedule'))


@login_required(login_url='organization:login')
def admin_redirect(request, **kwargs):
    if request.user.is_staff:
        if 'whereto' in kwargs:
            return HttpResponseRedirect(reverse_lazy('admin:index') + 'entry/' + kwargs['whereto'] + "/")
        return HttpResponseRedirect(reverse_lazy('admin:index'))
    return HttpResponseRedirect(reverse_lazy('entry:index'))


def handler404(request, exception, template_name="entry/secret.html"):
    print(exception)
    response = render(request, template_name)
    response.status_code = 404
    return response


class FRCdata(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry.login'
    template_name = 'entry/frc.html'

    def get(self, request, **kwargs):
        # get_team_logos()
        # import_first()
        # get_team_list()
        get_match_data_event(Event.objects.get(id=121))
        try:
            context = {

            "img":Team.objects.get(number=167).avatar
            }
        except:
            context = {

            "img":"NA"
            }
        return render(request, 'entry/frc.html', context)


class TeamSettingsNotFoundError(LoginRequiredMixin, generic.TemplateView):
    login_url = 'organization:login'
    template_name = 'entry/team_settings_not_found_error.html'


class TeamList(LoginRequiredMixin, generic.ListView):
    login_url = 'organization:login'
    template_name = 'entry/teams.html'
    context_object_name = "team_list"
    model = Team

    def get_queryset(self):
        teams = get_present_teams(self.request.user)
        if teams.count() == 1 and teams.first() == Team.objects.first():
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

        return teams

class Import(LoginRequiredMixin, FormMixin, generic.TemplateView):
    login_url = 'organization:login'
    template_name = 'entry/import.html'
    model = Team
    form_class = ImportForm
    success_url = 'entry:index'
    importing: bool = False

    @classmethod
    def post(cls, request, *args, **kwargs):
        if cls.importing:
            return render(request, 'entry/import.html', {'importing': True, 'form': ImportForm()})

        form = ImportForm(request.POST)
        context = {'form': form}
        success = False
        if form.is_valid():
            import_type = form.cleaned_data['import_type']
            key = form.cleaned_data['key']
            cls.importing = True

            if import_type == 1:  # Event key
                success = import_first(event_code=key)
            elif import_type == 2:  # Team Number key
                success = import_first(team_number=key)
            elif import_type == 3:  # Import all. TODO disable before release
                success = import_first()

        cls.importing = False
        context['importing'] = False
        context['success'] = success
        return render(request, 'entry/import.html', context)

    @classmethod
    def get(cls, request, *args, **kwargs):
        return render(request, 'entry/import.html', {'form': ImportForm(), 'importing': cls.importing})


class Index(LoginRequiredMixin, generic.TemplateView):
    login_url = 'organization:login'
    template_name = 'entry/index.html'
    model = Team


class Experimental(LoginRequiredMixin, generic.TemplateView):
    login_url = 'organization:login'
    model = Team
    template_name = 'entry/experimental.html'

    def get(self, request, *args, **kwargs):
        form = MatchScoutForm()
        rendered_form = form.render()
        context = {'form': rendered_form}
        return render(request, 'entry/experimental.html', context)


class About(LoginRequiredMixin, generic.TemplateView):
    login_url = 'organization:login'
    template_name = 'entry/about.html'


class Tutorial(generic.TemplateView):
    template_name = 'entry/tutorial.html'


class Welcome(LoginRequiredMixin, generic.TemplateView):
    template_name = 'entry/welcome.html'


class Upload(LoginRequiredMixin, generic.TemplateView):
    login_url = 'organization:login'
    template_name = 'entry/upload.html'


class Settings(LoginRequiredMixin, generic.TemplateView):
    login_url = 'organization:login'
    template_name = 'entry/settings.html'

    def post(self, request, *args, **kwargs):
        response = HttpResponseRedirect(reverse_lazy('entry:settings'))
        response.set_cookie('images', request.POST.get('images'))
        response.set_cookie('filters', request.POST.get('filters'))
        response.set_cookie('districtTeams', request.POST.get('districtTeams'))
        response.set_cookie('tutorialCompleted', request.POST.get('tutorialCompleted'))
        response.set_cookie('teamsBehaviour', request.POST.get('teamsBehaviour'))
        response.set_cookie('teamListType', request.POST.get('teamListType'))

        if self.request.user.orgmember.position == "LS":
            new_settings = OrgSettings()
            print("making")
            try:
                new_settings = OrgSettings.objects.get(organization=self.request.user.orgmember.organization)
                new_settings.current_event = Event.objects.get(FIRST_key=request.POST.get('currentEvent'))

            except OrgSettings.DoesNotExist:
                new_settings.organization = self.request.user.orgmember.organization
                new_settings.current_event = Event.objects.get(FIRST_key=request.POST.get('currentEvent'))

            except Event.DoesNotExist:
                print("User entered event that doesn't exist")
                return response

            new_settings.save()

        return response
