import os
import ast
import json

import requests
from openpyxl import Workbook
from datetime import datetime

from django.core import serializers
from django.shortcuts import render
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import HttpResponseRedirect, HttpResponse, Http404, JsonResponse
from django.views import generic
from django.views.decorators.csrf import csrf_exempt
from django.template import Library, loader
from django.contrib import auth
from django.contrib.auth.decorators import login_required
from django.views.generic.edit import FormMixin
from django_ajax.decorators import ajax
from django.conf import settings

from apps.entry.graphing import *
from apps.entry.templatetags.common_tags import *
from apps import importFRC
from apps.entry.forms import MatchScoutForm, RegistrationForm, PitScoutForm, LoginForm, ImportForm

register = Library


@ajax
@csrf_exempt
@login_required(login_url='entry:login')
def update_graph(request):
    graph_type = request.POST.getlist('graphType')[0]
    try:
        output = graph(graph_type, request)
        if output == "lazy":
            return HttpResponseRedirect(reverse_lazy('entry:visualize'))
        print(output)
        response = HttpResponse(json.dumps(output), content_type="application/json")
        return response
    except IOError:
        print("Image not found")
        return Http404
    except NoTeamsProvided:
        return NoTeamsProvided
    except NoFieldsProvided:
        return NoFieldsProvided


@ajax
@csrf_exempt
@login_required(login_url='entry:login')
def update_glance(request, pk):
    matches = Match.objects.filter(team_id=pk,
                                   ownership_id=request.user.orgmember.organization_id).order_by('event',
                                                                                                 'match_number')
    count = matches.count()
    try:
        if make_int(Team.objects.get(id=pk).glance.name.split('_')[2]) == count:
            return HttpResponse(Team.objects.get(id=pk).glance.read(), content_type='application/json')
    except IndexError:
        print("")
    except AttributeError:
        print("")
    except FileNotFoundError:
        print("Creating new glance json file for " + str(Team.objects.get(id=pk).glance))

    matches_json = serializers.serialize('json', matches)
    if not settings.USE_MEDIA_SPACES:
        f = open(os.path.join(settings.BASE_DIR, 'glance_temp.json'), 'w')
        f.write(str(matches_json))
        f = open(os.path.join(settings.BASE_DIR, 'glance_temp.json'), 'r', encoding='UTF-8')
        team = Team.objects.get(id=pk)
        team.glance.delete()
        team.glance.save(
            'glance_' + str(pk) + '_' + str(count) + '_' + str(datetime.datetime.now()) + '.json', f)
    return HttpResponse(matches_json, content_type='application/json')


@ajax
@csrf_exempt
@login_required(login_url='entry:login')
def update_fields(request):
    if request.method == "GET":
        try:
            path = 'scoring.json'
            path = os.path.join(settings.BASE_DIR, path)
            with open(path) as f:
                result = json.load(f)
            return JsonResponse(result)
        except IOError:
            print("Fields file not found")
            return Http404
    else:
        return HttpResponseRedirect(reverse_lazy('entry:visualize'))


def decode_ajax(request):
    return dict(QueryDict(request.body.decode()))


def scout_lead_check(user):
    return user.groups.filter(name="Scouting").exists()


# @user_passes_test(scout_lead_check, login_url='entry:login')
@login_required(login_url='entry:login')
def download(request):
    # TODO find a way to prevent spamming this.

    path = 'match_history.xlsx'
    path = os.path.join(settings.BASE_DIR, path)
    update_csv(request.user.orgmember.organization)

    if request.method == "GET" and os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="text/xlsx")
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(path)
            return response

    return Http404


@ajax
@csrf_exempt
@login_required(login_url='entry:login')
def get_csv_ajax(request):
    path = 'match_history.xlsx'
    path = os.path.join(settings.BASE_DIR, path)
    update_csv(request.user.orgmember.organization)

    if os.path.exists(path):
        with open(path, 'rb') as fh:
            response = HttpResponse(fh.read(), content_type="text/xlsx")
            response['Content-Disposition'] = 'attachment; filename=' + os.path.basename(path)
            return response
    return Http404


def update_csv(organization):
    print("Updating XLSX File")

    workbook = Workbook()
    writable_models = [Match, Pits]

    for model in writable_models:
        workbook.create_sheet(title=model._meta.model_name)
        sheet = workbook.get_sheet_by_name(model._meta.model_name)
        headers = model._meta.get_fields()
        x = 1
        for header in headers:
            sheet.cell(column=x, row=1, value=str(header))
            x += 1
        data = model.objects.all().values_list().filter(ownership=organization)
        x = 1
        y = 2
        for entry in data:
            for field in entry:
                sheet.cell(column=x, row=y, value=str(field))
                x += 1
            y += 1

    workbook.remove_sheet(workbook.get_sheet_by_name("Sheet"))
    workbook.save("match_history.xlsx")


@login_required(login_url='entry:login')
def write_image_upload(request):
    if request.method == 'POST':
        team_number = make_int(request.POST.get('teamNumber', 0))
        team = Team.objects.get(number=team_number)

        request.session.set_test_cookie()

        files = request.FILES
        files = files.popitem()[1]

        for file in files:
            file.name = str(team_number) + "-----" + str(datetime.datetime.now()).replace('.', '') + '.jpg'
            image = Images(name=team.name, image=file)
            image.save()
            team.images.add(image)
            team.save()
        return HttpResponseRedirect(reverse_lazy('entry:index'))
    else:
        return HttpResponseRedirect(reverse_lazy('entry:index'))


@login_required(login_url='entry:login')
def write_pit_upload(request):
    if request.method == 'GET':
        return HttpResponseRedirect(reverse_lazy('entry:index'))


@login_required(login_url='entry:login')
def import_from_first(request):
    if request.method == 'GET':
        return HttpResponseRedirect(reverse_lazy('entry:import'))

    elif request.method == 'POST':
        import_type = make_int(request.POST.get('importType', 0))
        key = request.POST.get('key', 0)
        year = request.POST.get('year', 0)

        if year == 0:
            year = None

        if import_type == 0:
            importFRC.import_district(key, year)
        elif import_type == 1:
            importFRC.import_event(key, year)
        elif import_type == 3:
            importFRC.import_team(key, year)

    return HttpResponseRedirect(reverse_lazy('entry:import'))


@login_required(login_url='entry:login')
def import_schedule_from_first(request):
    teamsettings = OrgSettings.objects.all().filter(org_id=request.user.orgmember.organization_id)[0]
    first_key = Event.objects.all().filter(id=make_int(teamsettings.current_event.id))[0].FIRST_key
    importFRC.import_schedule(first_key, playoffs=False)
    importFRC.import_schedule(first_key, playoffs=True)

    return HttpResponseRedirect(reverse_lazy('entry:schedule'))


@login_required(login_url='entry:login')
def logout(request):
    auth.logout(request)
    print("request.user.is_authenticated:" + str(request.user.is_authenticated))
    return HttpResponseRedirect(reverse_lazy('entry:index'))


@login_required(login_url='entry:login')
def admin_redirect(request, **kwargs):
    if request.user.is_staff:
        if 'whereto' in kwargs:
            return HttpResponseRedirect(reverse_lazy('admin:index') + 'entry/' + kwargs['whereto'] + "/")
        return HttpResponseRedirect(reverse_lazy('admin:index'))
    return HttpResponseRedirect(reverse_lazy('entry:index'))


def handler404(request, exception, template_name="entry/secret.html"):
    print(exception)
    response = render(request, template_name)
    response.status_code = 404
    return response


def make_int(s):
    if isinstance(s, str):
        if len(s) == 0:
            return 0
    if s == 'False':
        return False
    elif s == 'True':
        return True
    s = str(s)
    s = s.strip()
    return int(s) if s else 0


def get_present_teams(user):
    try:
        # TODO See if this is actually the best way to query all teams from attendance...
        # Note. We do need the actual Team objects (name, number, colour etc...) all that jazz
        objects = Team.objects.filter(
            number__in=Attendance.objects.filter(
                event=user.orgmember.organization.settings.current_event
            ).values_list('team_id', flat=True)
        )
        return objects
    except OrgSettings.DoesNotExist:
        return Team.objects.all()


# TODO Remove this
def get_all_teams():
    objects = Team.objects.all()
    objects = objects.order_by('number')
    return objects


# TODO Remove This
def get_all_events():
    return Event.objects.all().order_by('start')


def handle_query_present_teams(view):
    teams = get_present_teams(view.request.user)
    if teams.count() == 1 and teams.first() == Team.objects.first():
        return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

    return teams

class FRCdata(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry.login'
    template_name = 'entry/frc.html'
    def get(self,request):
        responce = requests.get('https://frc-api.firstinspires.org/v3.0/2023/teams', headers = {'Authorization': 'Basic dm9ydGV4MTQ4OmFkY2E5ZDI5LTU3YWUtNDJiMi1hMTY3LWZjMDhiMzg2Mzg4OQ=='}).json()
        # responce = responce[]
        print(responce)
        return render(request, 'entry/frc.html')

class TeamSettingsNotFoundError(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry:login'
    template_name = 'entry/team_settings_not_found_error.html'


class TeamList(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    template_name = 'entry/teams.html'
    context_object_name = "team_list"
    model = Team
    def get_queryset(self):
        teams = get_present_teams(self.request.user)
        if teams.count() == 1 and teams.first() == Team.objects.first():
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

        return teams

class Import(LoginRequiredMixin, FormMixin, generic.TemplateView):
    login_url = 'entry:login'
    template_name = 'entry/import.html'
    model = Team
    form_class = ImportForm
    success_url = 'entry:index'

    @staticmethod
    def post(request, *args, **kwargs):
        form = ImportForm(request.POST)
        context = {'form': form}
        if form.is_valid():
            import_type = form.cleaned_data['import_type']
            key = form.cleaned_data['key']
            year = '2022'

            if import_type == 0:
                importFRC.import_district(key, year)
            elif import_type == 1:
                importFRC.import_event(key, year)
            elif import_type == 2:
                importFRC.import_team(key, year)

        return render(request, 'entry/import.html', context)


class Index(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry:login'
    template_name = 'entry/index.html'
    model = Team


class MatchScout(LoginRequiredMixin, FormMixin, generic.DetailView):
    login_url = 'entry:login'
    model = Team
    template_name = 'entry/matchscout.html'
    form_class = MatchScoutForm
    success_url = 'entry:match_scout_landing'

    @staticmethod
    def post(request, pk, *args, **kwargs):
        org_settings = request.user.orgmember.organization.settings

        form = MatchScoutForm(request.POST,
                              event=org_settings.current_event,
                              ownership=request.user.orgmember.organization)
        team = Team.objects.get(id=pk)
        context = {'form': form, 'team': team}

        if form.is_valid():
            print("Valid Match Scout")
            print(form.cleaned_data)
            first_key = Event.objects.all().filter(id=make_int(org_settings.current_event.id))[0].FIRST_key

            match = Match(**form.cleaned_data)
            match.team = team
            match.event = Event.objects.get(FIRST_key=first_key)
            match.scouter_name = request.user.username
            match.ownership = request.user.orgmember.organization
            try:
                match.save()
                print('Match Scout Submission Success')
            except Exception as e:
                print(e)

            print(match)

            return HttpResponseRedirect(reverse_lazy('entry:match_scout_landing'))

        return render(request, 'entry/matchscout.html', context)


class MatchScoutLanding(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    model = Team
    context_object_name = "team_list"
    template_name = 'entry/matchlanding.html'

    def get_queryset(self):
        teams = get_present_teams(self.request.user)
        if teams.count() == 1 and teams.first() == Team.objects.first():
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

        return teams


class PitScout(LoginRequiredMixin, FormMixin, generic.DetailView):
    login_url = 'entry:login'
    template_name = 'entry/pitscout.html'
    model = Team
    context_object_name = "team"
    form_class = PitScoutForm
    success_url = 'entry:pit_scout_landing'

    @staticmethod
    def post(request, pk, *args, **kwargs):
        form = PitScoutForm(request.POST)
        team = Team.objects.get(id=pk)
        context = {'form': form, 'team': team}

        if form.is_valid():
            pits = Pits(**form.cleaned_data)
            # TODO Finish This
            return HttpResponseRedirect(reverse_lazy('entry:pit_scout_landing'))
        return render(request, 'entry/pitscout.html', context)


class PitScoutLanding(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    template_name = 'entry/pitlanding.html'
    context_object_name = "team_list"

    def get_queryset(self):
        teams = get_present_teams(self.request.user)
        if teams.count() == 1 and teams.first() == Team.objects.first():
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

        return teams


class Visualize(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    template_name = 'entry/visualization.html'
    model = Team
    context_object_name = "team_list"

    def get_queryset(self):
        teams = get_present_teams(self.request.user)
        if teams.count() == 1 and teams.first() == Team.objects.first():
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))
        return teams


class ScheduleView(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    template_name = 'entry/schedule.html'
    context_object_name = "schedule_list"
    model = Schedule
    show_completed = False

    def get_queryset(self):
        try:
            org_settings = self.request.user.orgmember.organization.settings
        except IndexError as e:
            print(str(e) + ": There are no team settings for this query.")
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

        schedule = Schedule.objects.filter(event_id=org_settings.current_event)\
                    .exclude(match_number=0)\
                    .order_by("match_type")\
                    .order_by("match_number")
        results = Result.objects \
            .filter(event=org_settings.current_event, ownership=org_settings.organization, completed=True) \
            .values_list('match__match_number', flat=True)

        if self.show_completed:
            return schedule, results, True

        return schedule, results, False


class ScheduleDetails(LoginRequiredMixin, generic.DetailView):
    login_url = 'entry:login'
    template_name = 'entry/schedule-details.html'
    model = Schedule
    context_object_name = "schedule"


class Experimental(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry:login'
    model = Team
    template_name = 'entry/experimental.html'

    def get(self, request, *args, **kwargs):
        form = MatchScoutForm()
        rendered_form = form.render()
        context = {'form': rendered_form}
        return render(request, 'entry/experimental.html', context)


class About(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry:login'
    template_name = 'entry/about.html'


class Tutorial(generic.TemplateView):
    template_name = 'entry/tutorial.html'


class Welcome(LoginRequiredMixin, generic.TemplateView):
    template_name = 'entry/welcome.html'


class Glance(LoginRequiredMixin, generic.DetailView):
    login_url = 'entry:login'
    model = Team
    template_name = 'entry/glance.html'
    context_object_name = "team"

    def head(self, *args, **kwargs):
        output = {
            "test": 1
        }
        response = HttpResponse(json.dumps(output), content_type="application/json")
        return response


class GlanceLanding(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    model = Team
    template_name = 'entry/glancelanding.html'

    def get_queryset(self):
        return handle_query_present_teams(self)


class Login(FormMixin, generic.TemplateView):
    template_name = 'entry/login.html'
    form_class = LoginForm
    success_url = 'entry:index'

    def post(self, request, *args, **kwargs):
        form = LoginForm(request.POST)
        context = {'form': form}
        if form.is_valid():
            user = auth.authenticate(
                request,
                username=form.cleaned_data['username'],
                password=form.cleaned_data['password'])
            if user is not None:
                auth.login(request, user)
                if not OrgMember.objects.filter(user=user).exists():
                    OrgMember.objects.create(user_id=user.id)
            return HttpResponseRedirect(reverse_lazy('entry:index'))

        form.add_error('username', 'Username or password is incorrect')
        form.add_error('password', 'Username or password is incorrect')
        context = {'form': form}
        return render(request, 'entry/login.html', context)


class Registration(FormMixin, generic.TemplateView):
    template_name = 'entry/register.html'
    form_class = RegistrationForm
    success_url = 'entry:index'

    def get(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return HttpResponseRedirect(reverse_lazy('entry:index'))
        else:
            return super(Registration, self).get(request, *args, **kwargs)

    @staticmethod
    def post(request, *args, **kwargs):
        form = RegistrationForm(request.POST)
        context = {'form': form}
        if form.is_valid():
            user = User()
            user.orgmember = OrgMember()
            user.set_password(form.cleaned_data['password'])
            user.username = form.cleaned_data['username']
            user.first_name = form.cleaned_data['first_name']
            user.last_name = form.cleaned_data['last_name']
            user.email = form.cleaned_data['email']
            print(user)
            try:
                if form.cleaned_data['create_new_org']:
                    org = Organization()
                    org.settings = OrgSettings()
                    org.settings.current_event = Event.objects.first()
                    org.name = form.cleaned_data['org_name']
                    user.orgmember.position = 'LS'
                    user.orgmember.organization = org
                else:
                    org = Organization.objects.get(name=form.cleaned_data['org_name'])
                    if str(org.reg_id)[:6] != form.cleaned_data['org_reg_id']:
                        raise Organization.DoesNotExist
                    user.orgmember.organization = org
            except Organization.DoesNotExist:
                form.add_error('org_name', "Org Name or UUID is incorrect.")
                form.add_error('org_reg_id', "Org Name or UUID is incorrect.")
                context = {'form': form}
                print('org does not exist ' + form.cleaned_data)
                return render(request, 'entry/register.html', context)

            print(user)
            user.save()
            user.orgmember.organization.settings.save()
            user.orgmember.organization.save()
            user.orgmember.save()
            user.save()

            user = auth.authenticate(username=form.cleaned_data['username'], password=form.cleaned_data['password'])
            print(user)
            auth.login(request, user)
            return HttpResponseRedirect(reverse_lazy('entry:index'))

        return render(request, 'entry/register.html', context)


class MatchData(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    template_name = 'entry/matchdata.html'
    model = Match

    def get_queryset(self):
        try:
            org_settings = self.request.user.orgmember.organization.settings
        except IndexError:
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))
        return Match.objects.all().filter(event_id=org_settings.current_event).filter(
            ownership=self.request.user.orgmember.organization_id)


class PitData(LoginRequiredMixin, generic.ListView):
    login_url = 'entry:login'
    template_name = 'entry/pitdata.html'
    model = Pits

    def get_queryset(self):
        try:
            org_settings = self.request.user.orgmember.organization.settings
        except IndexError:
            return HttpResponseRedirect(reverse_lazy('entry:team_settings_not_found_error'))

        return Pits.objects.all().filter(event_id=org_settings.current_event).filter(
            ownership=self.request.user.orgmember.organization_id)


class Upload(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry:login'
    template_name = 'entry/upload.html'


class Settings(LoginRequiredMixin, generic.TemplateView):
    login_url = 'entry:login'
    template_name = 'entry/settings.html'

    def post(self, request, *args, **kwargs):
        response = HttpResponseRedirect(reverse_lazy('entry:settings'))
        response.set_cookie('images', request.POST.get('images', ''))
        response.set_cookie('filters', request.POST.get('filters', ''))
        response.set_cookie('districtTeams', request.POST.get('districtTeams', ''))
        response.set_cookie('tutorialCompleted', request.POST.get('tutorialCompleted', ''))
        response.set_cookie('teamsBehaviour', request.POST.get('teamsBehaviour', ''))
        response.set_cookie('teamListType', request.POST.get('teamListType', ''))

        if self.request.user.orgmember.position == "LS":
            new_settings = OrgSettings()
            print("making")
            try:
                new_settings = OrgSettings.objects.get(organization=self.request.user.orgmember.organization)
                new_settings.current_event = Event.objects.get(FIRST_key=request.POST.get('currentEvent', '21ONT'))

            except OrgSettings.DoesNotExist:
                new_settings.organization = self.request.user.orgmember.organization
                new_settings.current_event = Event.objects.get(FIRST_key=request.POST.get('currentEvent', '21ONT'))

            except Event.DoesNotExist:
                print("User entered event that doesn't exist")
                return response

            new_settings.save()

        return response
